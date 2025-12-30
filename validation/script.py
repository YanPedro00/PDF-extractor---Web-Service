"""
Script para extração minuciosa de textos em PDFs escaneados
Detecta espaçamentos exatos entre palavras usando OCR + Bounding Boxes
Usa Tesseract OCR para maior precisão e simplicidade
Valida contra ground truth para medir precisão
"""

import os
import json
import re
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def exportar_para_excel(dados_extraidos, nome_arquivo="dados_extraidos.xlsx"):
    """
    Exporta os dados extraídos do PDF para um arquivo Excel formatado.
    Gera apenas as informações relevantes do PDF (sem dados técnicos da análise).
    
    Args:
        dados_extraidos: Lista de dicionários com dados extraídos
        nome_arquivo: Nome do arquivo Excel de saída
    
    Returns:
        str: Caminho do arquivo Excel gerado
    """
    if not dados_extraidos:
        print("⚠️  Nenhum dado para exportar.")
        return None
    
    print(f"\n{'=' * 100}")
    print("📊 EXPORTANDO DADOS PARA EXCEL")
    print(f"{'=' * 100}\n")
    
    # Preparar dados para DataFrame - APENAS INFORMAÇÕES RELEVANTES
    dados_formatados = []
    
    for item in dados_extraidos:
        codigo_info = item.get('codigo', {})
        texto_completo = item.get('texto_linha', '')
        
        # Montar código completo preservando espaços
        codigo_base = codigo_info.get('codigo_base', '')
        sufixo = codigo_info.get('sufixo', '')
        num_espacos = codigo_info.get('espacos_detectados', 1)
        
        # Reconstruir código com espaços corretos
        codigo_completo = f"{codigo_base}{' ' * num_espacos}{sufixo}"
        
        # Extrair descrição e valor do texto
        resto_texto = texto_completo.replace(codigo_base, '', 1)
        resto_texto = resto_texto.replace(sufixo, '', 1).strip()
        
        # Separar descrição e valor
        partes = resto_texto.split('R$')
        if len(partes) >= 2:
            descricao = partes[0].strip()
            valor = f"R$ {partes[1].strip()}"
        else:
            descricao = resto_texto
            valor = ""
        
        # Adicionar apenas colunas relevantes do PDF
        dados_formatados.append({
            'Código': codigo_completo,
            'Descrição Item': descricao,
            'Valor': valor
        })
    
    # Criar DataFrame
    df = pd.DataFrame(dados_formatados)
    
    # Salvar no Excel
    df.to_excel(nome_arquivo, index=False, engine='openpyxl')
    
    # Aplicar formatação
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    ws.title = "Dados PDF"
    
    # Estilo do cabeçalho (cinza mais neutro)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formatação ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Ajustar largura das colunas
    column_widths = {
        'A': 20,  # Código
        'B': 40,  # Descrição Item
        'C': 15   # Valor
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Formatar células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Centralizar coluna de valor
            if cell.column_letter == 'C':
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adicionar informações no rodapé (opcional, discreto)
    footer_row = ws.max_row + 2
    ws[f'A{footer_row}'] = f"Extraído em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f'A{footer_row}'].font = Font(italic=True, size=9, color="808080")
    
    ws[f'C{footer_row}'] = f"Total: {len(dados_formatados)} registros"
    ws[f'C{footer_row}'].font = Font(italic=True, size=9, color="808080")
    ws[f'C{footer_row}'].alignment = Alignment(horizontal="right")
    
    # Salvar
    wb.save(nome_arquivo)
    
    print(f"✅ Arquivo Excel criado: {nome_arquivo}")
    print(f"📊 Total de registros: {len(dados_formatados)}")
    print(f"📋 Colunas: Código | Descrição Item | Valor")
    print(f"\n{'=' * 100}\n")
    
    return nome_arquivo


def carregar_ground_truth(json_path="ground_truth.json"):
    """
    Carrega o ground truth do arquivo JSON.
    
    Args:
        json_path: Caminho para o arquivo JSON com ground truth
    
    Returns:
        dict: Dados do ground truth ou None se não encontrado
    """
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def detectar_codigo_com_espacos_por_bbox(palavras_linha):
    """
    Detecta códigos no padrão XXXXX-XXXX [espaços] XA usando bounding boxes.
    Mais preciso que regex pois analisa distâncias reais em pixels.
    
    Args:
        palavras_linha: Lista de dicionários com informações das palavras
    
    Returns:
        dict ou None: Informações do código detectado ou None
    """
    # Procurar padrão: código-base seguido de sufixo XA
    padrao_codigo = r'\d{5}-\d{4}'
    padrao_sufixo = r'\d+A'
    
    for i, palavra in enumerate(palavras_linha):
        if re.match(padrao_codigo, palavra['text']):
            # Encontrou código base, verificar se próxima palavra é sufixo
            if i + 1 < len(palavras_linha):
                proxima = palavras_linha[i + 1]
                if re.match(padrao_sufixo, proxima['text']):
                    # Calcular espaços baseado em distância pixel
                    fim_codigo = palavra['left'] + palavra['width']
                    inicio_sufixo = proxima['left']
                    distancia = inicio_sufixo - fim_codigo
                    
                    # Estimar largura de um caractere do código
                    largura_char = palavra['width'] / len(palavra['text'])
                    
                    # Calcular número de espaços
                    # Um espaço tipicamente tem ~0.6 da largura de um caractere em fontes mono
                    # Ajustado empiricamente com base no ground truth
                    largura_espaco = largura_char * 0.6
                    num_espacos = round(distancia / largura_espaco)
                    
                    # Garantir pelo menos 1 espaço
                    num_espacos = max(1, num_espacos)
                    
                    return {
                        'codigo_base': palavra['text'],
                        'sufixo': proxima['text'],
                        'espacos_detectados': num_espacos,
                        'distancia_px': distancia,
                        'largura_char': largura_char
                    }
    
    return None


def detectar_codigo_com_espacos(texto):
    """
    Detecta códigos no padrão XXXXX-XXXX [espaços] XA no texto (regex).
    NOTA: Menos preciso que análise por bounding box.
    
    Args:
        texto: String de texto para analisar
    
    Returns:
        dict ou None: Informações do código detectado ou None
    """
    # Padrão: 5 dígitos - 4 dígitos [espaços] dígito+A
    padrao = r'(\d{5}-\d{4})(\s+)(\d+A)'
    match = re.search(padrao, texto)
    
    if match:
        codigo_base = match.group(1)
        espacos = match.group(2)
        sufixo = match.group(3)
        num_espacos = len(espacos)
        
        return {
            'codigo_completo': f"{codigo_base}{espacos}{sufixo}",
            'codigo_base': codigo_base,
            'sufixo': sufixo,
            'espacos_detectados': num_espacos
        }
    
    return None


def calcular_espacos_entre_palavras(palavra1_data, palavra2_data):
    """
    Calcula o número de espaços entre duas palavras baseado em suas posições.
    
    Args:
        palavra1_data: dict com dados da primeira palavra {'text', 'left', 'width'}
        palavra2_data: dict com dados da segunda palavra {'text', 'left', 'width'}
    
    Returns:
        int: Número estimado de espaços
    """
    # Posição onde termina a primeira palavra
    fim_palavra1 = palavra1_data['left'] + palavra1_data['width']
    
    # Posição onde começa a segunda palavra
    inicio_palavra2 = palavra2_data['left']
    
    # Distância entre as palavras
    distancia = inicio_palavra2 - fim_palavra1
    
    # Estimar largura média de um caractere da primeira palavra
    if len(palavra1_data['text']) > 0:
        largura_char = palavra1_data['width'] / len(palavra1_data['text'])
    else:
        largura_char = 10  # fallback
    
    # Largura de um espaço é tipicamente 0.3-0.5 da largura de um caractere
    # Vamos usar 0.4 como padrão
    largura_espaco = largura_char * 0.4
    
    if largura_espaco == 0:
        return 1
    
    # Calcular número de espaços
    num_espacos = round(distancia / largura_espaco)
    
    # Garantir pelo menos 1 espaço entre palavras
    return max(1, num_espacos)


def processar_linha(palavras_linha):
    """
    Processa uma linha de palavras, detectando espaçamentos entre elas.
    
    Args:
        palavras_linha: Lista de dicionários com informações das palavras
    
    Returns:
        str: Texto reconstruído com espaçamentos detectados
        list: Informações detalhadas sobre espaçamentos
    """
    if not palavras_linha:
        return "", []
    
    texto_completo = ""
    infos_espacamento = []
    
    for i, palavra_data in enumerate(palavras_linha):
        texto_completo += palavra_data['text']
        
        # Se não for a última palavra da linha
        if i < len(palavras_linha) - 1:
            proxima_palavra = palavras_linha[i + 1]
            
            # Calcular número de espaços
            num_espacos = calcular_espacos_entre_palavras(palavra_data, proxima_palavra)
            
            # Adicionar espaços ao texto
            espacos_str = ' ' * num_espacos
            texto_completo += espacos_str
            
            # Registrar informação detalhada
            info = {
                'palavra1': palavra_data['text'],
                'palavra2': proxima_palavra['text'],
                'num_espacos': num_espacos,
                'distancia_px': proxima_palavra['left'] - (palavra_data['left'] + palavra_data['width']),
                'pos_palavra1': palavra_data['left'],
                'pos_palavra2': proxima_palavra['left']
            }
            infos_espacamento.append(info)
    
    return texto_completo, infos_espacamento


def agrupar_por_linhas(dados_ocr, tolerancia_y=5):
    """
    Agrupa palavras por linha baseado na posição vertical (y).
    
    Args:
        dados_ocr: DataFrame do Tesseract com dados de OCR
        tolerancia_y: Tolerância em pixels para considerar mesma linha
    
    Returns:
        dict: Dicionário com número da linha e lista de palavras
    """
    linhas = {}
    linha_atual = 0
    y_ref = None
    
    for i, row in dados_ocr.iterrows():
        # Ignorar campos vazios
        if not str(row['text']).strip() or row['conf'] < 0:
            continue
        
        # Só processar palavras (level 5)
        if row['level'] != 5:
            continue
        
        y_atual = row['top']
        
        # Determinar se é uma nova linha
        if y_ref is None:
            y_ref = y_atual
        elif abs(y_atual - y_ref) > tolerancia_y:
            linha_atual += 1
            y_ref = y_atual
        
        # Adicionar palavra à linha
        if linha_atual not in linhas:
            linhas[linha_atual] = []
        
        palavra_info = {
            'text': str(row['text']).strip(),
            'left': row['left'],
            'top': row['top'],
            'width': row['width'],
            'height': row['height'],
            'conf': row['conf']
        }
        linhas[linha_atual].append(palavra_info)
    
    # Ordenar palavras em cada linha pela posição horizontal (left)
    for linha_num in linhas:
        linhas[linha_num] = sorted(linhas[linha_num], key=lambda x: x['left'])
    
    return linhas


def extrair_pdf_com_espacamento(pdf_path, ground_truth_path="ground_truth.json"):
    """
    Função principal: Extrai texto de PDF escaneado detectando espaçamentos exatos.
    Valida contra ground truth se disponível.
    
    Args:
        pdf_path: Caminho para o arquivo PDF
        ground_truth_path: Caminho para o arquivo JSON com ground truth
    """
    print("=" * 100)
    print("🔍 EXTRAÇÃO MINUCIOSA DE PDF COM DETECÇÃO DE ESPAÇAMENTOS + VALIDAÇÃO")
    print("=" * 100)
    print(f"\n📄 Arquivo: {pdf_path}\n")
    
    # Carregar ground truth
    ground_truth = carregar_ground_truth(ground_truth_path)
    if ground_truth:
        print(f"✅ Ground truth carregado: {len(ground_truth['linhas'])} códigos esperados\n")
    else:
        print(f"⚠️  Ground truth não encontrado em: {ground_truth_path}\n")
    
    # Estrutura para armazenar resultados
    codigos_extraidos = []
    resultados_validacao = []
    
    # 1. Converter PDF para imagem
    print("⏳ Convertendo PDF para imagem (DPI 300 para alta qualidade)...")
    try:
        imagens = convert_from_path(pdf_path, dpi=300)
        print(f"✅ {len(imagens)} página(s) convertida(s)\n")
    except Exception as e:
        print(f"❌ Erro ao converter PDF: {e}")
        return
    
    # 2. Processar cada página
    for idx_pagina, imagem in enumerate(imagens, 1):
        print(f"\n{'=' * 90}")
        print(f"📄 PÁGINA {idx_pagina}")
        print(f"{'=' * 90}\n")
        
        print("⏳ Executando OCR com Tesseract (detectando bounding boxes)...")
        
        # Configuração do Tesseract para português e máxima qualidade
        config_tesseract = '--psm 6 -l por'  # PSM 6: assume um bloco uniforme de texto
        
        # Executar OCR e obter dados detalhados (palavra por palavra com posições)
        dados_ocr = pytesseract.image_to_data(
            imagem, 
            lang='por',
            config=config_tesseract,
            output_type=pytesseract.Output.DATAFRAME
        )
        
        # Filtrar apenas resultados com confiança razoável
        dados_ocr = dados_ocr[dados_ocr['conf'] > 30]
        
        if dados_ocr.empty:
            print("⚠️  Nenhum texto detectado com confiança suficiente.\n")
            continue
        
        print(f"✅ OCR concluído! {len(dados_ocr[dados_ocr['level'] == 5])} palavras detectadas\n")
        
        # Agrupar palavras por linhas
        linhas = agrupar_por_linhas(dados_ocr)
        
        print(f"📝 Total de {len(linhas)} linha(s) identificada(s)\n")
        print("-" * 90)
        
        # Processar cada linha
        for num_linha, palavras in sorted(linhas.items()):
            texto_linha, infos = processar_linha(palavras)
            
            # Tentar detectar código na linha usando bounding boxes (mais preciso)
            codigo_detectado = detectar_codigo_com_espacos_por_bbox(palavras)
            
            # Mostrar linhas com códigos detectados
            if codigo_detectado:
                print(f"\n{'=' * 100}")
                print(f"🎯 CÓDIGO DETECTADO NA LINHA {num_linha + 1}")
                print(f"{'=' * 100}")
                
                print(f"\n📝 Texto completo da linha:")
                print(f"   '{texto_linha}'")
                
                print(f"\n🔍 Análise do código (via Bounding Boxes):")
                print(f"   • Código base: {codigo_detectado['codigo_base']}")
                print(f"   • Sufixo: {codigo_detectado['sufixo']}")
                print(f"   • Espaços detectados: {codigo_detectado['espacos_detectados']}")
                print(f"   • Distância horizontal: {codigo_detectado['distancia_px']:.2f}px")
                print(f"   • Largura média/caractere: {codigo_detectado['largura_char']:.2f}px")
                
                espacos_visual = '·' * codigo_detectado['espacos_detectados']
                print(f"   • Visualização: {codigo_detectado['codigo_base']}{espacos_visual}{codigo_detectado['sufixo']}")
                
                # Armazenar para validação e exportação posterior
                codigos_extraidos.append({
                    'linha': num_linha + 1,
                    'codigo': codigo_detectado,
                    'texto_linha': texto_linha
                })
                
                # Validar contra ground truth se disponível
                if ground_truth:
                    # Procurar correspondência no ground truth
                    gt_linha = next((l for l in ground_truth['linhas'] 
                                   if l['codigo_base'] == codigo_detectado['codigo_base']), None)
                    
                    if gt_linha:
                        espacos_corretos = gt_linha['espacos_no_codigo']
                        espacos_detectados = codigo_detectado['espacos_detectados']
                        
                        if espacos_corretos == espacos_detectados:
                            status = "✅ CORRETO"
                            acerto = True
                        else:
                            status = "❌ INCORRETO"
                            acerto = False
                        
                        print(f"\n📊 VALIDAÇÃO vs GROUND TRUTH:")
                        print(f"   • Esperado: {espacos_corretos} espaço(s)")
                        print(f"   • Detectado: {espacos_detectados} espaço(s)")
                        print(f"   • Status: {status}")
                        
                        resultados_validacao.append({
                            'linha': num_linha + 1,
                            'codigo_base': codigo_detectado['codigo_base'],
                            'esperado': espacos_corretos,
                            'detectado': espacos_detectados,
                            'acerto': acerto
                        })
                
                print(f"\n{'-' * 100}")
    
    # Mostrar sumário de validação
    if resultados_validacao:
        print("\n" + "=" * 100)
        print("📊 SUMÁRIO DE VALIDAÇÃO")
        print("=" * 100)
        
        total = len(resultados_validacao)
        acertos = sum(1 for r in resultados_validacao if r['acerto'])
        erros = total - acertos
        precisao = (acertos / total * 100) if total > 0 else 0
        
        print(f"\n✅ Total de códigos validados: {total}")
        print(f"✅ Acertos: {acertos}")
        print(f"❌ Erros: {erros}")
        print(f"📈 Precisão: {precisao:.1f}%")
        
        if erros > 0:
            print(f"\n❌ Códigos com erro:")
            for r in resultados_validacao:
                if not r['acerto']:
                    print(f"   • Linha {r['linha']}: {r['codigo_base']}")
                    print(f"     - Esperado: {r['esperado']} espaço(s)")
                    print(f"     - Detectado: {r['detectado']} espaço(s)")
                    print(f"     - Diferença: {abs(r['esperado'] - r['detectado'])} espaço(s)")
        
        print("\n" + "=" * 100)
    
    print("\n" + "=" * 100)
    print("✅ EXTRAÇÃO CONCLUÍDA COM SUCESSO!")
    print("=" * 100)
    
    # Exportar dados para Excel
    if codigos_extraidos:
        arquivo_excel = exportar_para_excel(codigos_extraidos, "dados_extraidos.xlsx")
    
    return codigos_extraidos, resultados_validacao


if __name__ == "__main__":
    # Testar com o PDF gerado
    pdf_teste = "matriz-ES-2025-2.pdf"
    
    if os.path.exists(pdf_teste):
        extrair_pdf_com_espacamento(pdf_teste)
    else:
        print(f"❌ Erro: Arquivo '{pdf_teste}' não encontrado.")
        print(f"   Diretório atual: {os.getcwd()}")
        print(f"   Arquivos disponíveis: {os.listdir('.')}")
