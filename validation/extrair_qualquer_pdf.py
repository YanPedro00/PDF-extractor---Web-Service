"""
Script genérico para extração de QUALQUER PDF com tabelas
Extrai todas as linhas e colunas automaticamente
Exporta para Excel formatado
"""

import os
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import re


def extrair_tabela_generica(pdf_path, output_excel="resultado_extracao.xlsx"):
    """
    Extrai TODA a tabela de um PDF de forma genérica.
    
    Args:
        pdf_path: Caminho do PDF
        output_excel: Nome do arquivo Excel de saída
    
    Returns:
        DataFrame com os dados extraídos
    """
    print("=" * 100)
    print("🔍 EXTRAÇÃO GENÉRICA DE PDF - QUALQUER TIPO DE TABELA")
    print("=" * 100)
    print(f"\n📄 Arquivo: {pdf_path}\n")
    
    if not os.path.exists(pdf_path):
        print(f"❌ Erro: Arquivo '{pdf_path}' não encontrado!")
        return None
    
    # 1. Converter PDF para imagem
    print("⏳ Convertendo PDF para imagem (DPI 300)...")
    imagens = convert_from_path(pdf_path, dpi=300)
    print(f"✅ {len(imagens)} página(s) convertida(s)\n")
    
    # 2. Extrair dados de todas as páginas
    todas_linhas = []
    
    for idx_pagina, imagem in enumerate(imagens, 1):
        print(f"\n{'=' * 100}")
        print(f"📄 PÁGINA {idx_pagina}")
        print(f"{'=' * 100}\n")
        
        print("⏳ Executando OCR...")
        
        # OCR com dados estruturados
        dados_ocr = pytesseract.image_to_data(
            imagem,
            lang='por',
            output_type=pytesseract.Output.DATAFRAME
        )
        
        # Filtrar apenas palavras com boa confiança
        dados_ocr = dados_ocr[dados_ocr['conf'] > 30]
        
        print(f"✅ {len(dados_ocr)} palavras detectadas")
        
        # Agrupar por linhas (baseado em posição Y)
        linhas_pagina = agrupar_por_linhas(dados_ocr)
        
        print(f"📝 {len(linhas_pagina)} linha(s) identificada(s)\n")
        
        # Adicionar todas as linhas
        for num_linha, palavras in sorted(linhas_pagina.items()):
            # Ordenar palavras por posição X
            palavras_ordenadas = sorted(palavras, key=lambda x: x['left'])
            
            # Montar texto da linha
            texto_linha = ' '.join([p['text'] for p in palavras_ordenadas])
            
            if texto_linha.strip():
                todas_linhas.append({
                    'pagina': idx_pagina,
                    'linha': num_linha + 1,
                    'texto': texto_linha.strip()
                })
    
    # 3. Criar DataFrame
    if not todas_linhas:
        print("⚠️  Nenhum dado extraído do PDF.")
        return None
    
    df = pd.DataFrame(todas_linhas)
    
    print(f"\n{'=' * 100}")
    print(f"📊 DADOS EXTRAÍDOS")
    print(f"{'=' * 100}\n")
    print(f"✅ Total de linhas extraídas: {len(df)}")
    print(f"✅ Total de páginas: {df['pagina'].nunique()}")
    
    # 4. Exportar para Excel
    print(f"\n⏳ Exportando para Excel: {output_excel}...")
    
    # Criar Excel formatado
    df_export = df[['texto']].copy()
    df_export.columns = ['Texto Extraído']
    
    df_export.to_excel(output_excel, index=False, engine='openpyxl')
    
    # Aplicar formatação
    wb = load_workbook(output_excel)
    ws = wb.active
    ws.title = "Dados Extraídos"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Bordas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aplicar formatação ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Largura da coluna
    ws.column_dimensions['A'].width = 120
    
    # Formatar células de dados
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Rodapé
    footer_row = ws.max_row + 2
    ws[f'A{footer_row}'] = f"Extraído em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Total: {len(df)} linhas"
    ws[f'A{footer_row}'].font = Font(italic=True, size=9, color="808080")
    
    wb.save(output_excel)
    
    print(f"✅ Excel criado: {output_excel}")
    
    # 5. Mostrar preview
    print(f"\n{'=' * 100}")
    print("📋 PREVIEW DOS DADOS (Primeiras 10 linhas)")
    print(f"{'=' * 100}\n")
    
    for idx, row in df.head(10).iterrows():
        print(f"[Pág {row['pagina']}] {row['texto'][:80]}...")
    
    print(f"\n{'=' * 100}")
    print("✅ EXTRAÇÃO CONCLUÍDA!")
    print(f"{'=' * 100}\n")
    
    return df


def agrupar_por_linhas(dados_ocr, tolerancia_y=5):
    """
    Agrupa palavras por linha baseado na posição Y.
    """
    linhas = {}
    linha_atual = 0
    y_ref = None
    
    for i, row in dados_ocr.iterrows():
        if not str(row['text']).strip() or row['conf'] < 0:
            continue
        
        if row['level'] != 5:  # Só processar palavras
            continue
        
        y_atual = row['top']
        
        # Determinar se é uma nova linha
        if y_ref is None:
            y_ref = y_atual
        elif abs(y_atual - y_ref) > tolerancia_y:
            linha_atual += 1
            y_ref = y_atual
        
        # Adicionar palavra à linha
        if linha_atual not in linhas:
            linhas[linha_atual] = []
        
        palavra_info = {
            'text': str(row['text']).strip(),
            'left': row['left'],
            'top': row['top'],
            'width': row['width'],
            'height': row['height'],
            'conf': row['conf']
        }
        linhas[linha_atual].append(palavra_info)
    
    return linhas


if __name__ == "__main__":
    # PDF a ser processado
    pdf_arquivo = "matriz-ES-2025-2.pdf"
    
    # Executar extração
    df = extrair_tabela_generica(pdf_arquivo, "matriz_extraida.xlsx")
    
    if df is not None:
        print(f"\n✅ Pronto! Abra o arquivo 'matriz_extraida.xlsx' para ver os dados.")

