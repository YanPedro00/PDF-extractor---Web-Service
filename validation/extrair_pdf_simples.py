"""
Script SIMPLES para extrair QUALQUER PDF com tabelas
Usa img2table + PaddleOCR para extração automática
Exporta para Excel formatado
"""

import os
import ssl
import re
import pandas as pd
from img2table.document import PDF
from img2table.ocr import PaddleOCR
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Correção SSL para Mac
ssl._create_default_https_context = ssl._create_unverified_context


def limpar_caracteres_invalidos(texto):
    """
    Remove caracteres inválidos para XML/Excel.
    """
    if not isinstance(texto, str):
        texto = str(texto) if texto is not None else ""
    # Manter apenas caracteres seguros para XML
    return re.sub(r'[^\x09\x0A\x0D\x20-\x7E\xA0-\xFF]', '', texto)


def extrair_pdf_para_excel(pdf_path, output_excel="dados_extraidos.xlsx"):
    """
    Extrai tabelas de qualquer PDF e exporta para Excel formatado.
    
    Args:
        pdf_path: Caminho do arquivo PDF
        output_excel: Nome do arquivo Excel de saída
    
    Returns:
        bool: True se sucesso, False caso contrário
    """
    print("=" * 100)
    print("🔍 EXTRAÇÃO AUTOMÁTICA DE PDF PARA EXCEL")
    print("=" * 100)
    print(f"\n📄 Arquivo: {pdf_path}\n")
    
    if not os.path.exists(pdf_path):
        print(f"❌ Erro: Arquivo '{pdf_path}' não encontrado!")
        return False
    
    try:
        # 1. Inicializar OCR
        print("⏳ Inicializando PaddleOCR...")
        ocr = PaddleOCR(lang='pt')
        print("✅ OCR pronto\n")
        
        # 2. Carregar PDF
        print("⏳ Carregando PDF...")
        doc = PDF(src=pdf_path)
        print("✅ PDF carregado\n")
        
        # 3. Extrair tabelas
        print("⏳ Extraindo tabelas (isso pode levar alguns segundos)...")
        tabelas = doc.extract_tables(
            ocr=ocr,
            implicit_rows=True,
            borderless_tables=True,
            min_confidence=40
        )
        
        if not tabelas:
            print("⚠️  Nenhuma tabela detectada no PDF.")
            return False
        
        print(f"✅ {len(tabelas)} página(s) com tabelas detectadas\n")
        
        # 4. Consolidar todas as tabelas em um único DataFrame
        todos_dados = []
        
        for pagina_idx, tabelas_pagina in tabelas.items():
            print(f"📄 Página {pagina_idx + 1}: {len(tabelas_pagina)} tabela(s)")
            
            for tab_idx, tabela in enumerate(tabelas_pagina):
                df = tabela.df
                
                # Adicionar coluna de identificação
                df.insert(0, 'Página', pagina_idx + 1)
                df.insert(1, 'Tabela', tab_idx + 1)
                
                todos_dados.append(df)
                print(f"  ├─ Tabela {tab_idx + 1}: {len(df)} linhas x {len(df.columns)} colunas")
        
        if not todos_dados:
            print("\n⚠️  Nenhum dado extraído.")
            return False
        
        # 5. Combinar todos os dados
        print(f"\n⏳ Consolidando dados...")
        df_final = pd.concat(todos_dados, ignore_index=True)
        
        print(f"✅ Total consolidado: {len(df_final)} linhas")
        
        # 5.5. Limpar caracteres inválidos
        print("⏳ Limpando caracteres inválidos...")
        df_final = df_final.map(limpar_caracteres_invalidos)
        
        # 6. Exportar para Excel
        print(f"⏳ Exportando para Excel: {output_excel}...")
        
        df_final.to_excel(output_excel, index=False, engine='openpyxl')
        
        # 7. Aplicar formatação
        print("⏳ Aplicando formatação...")
        wb = load_workbook(output_excel)
        ws = wb.active
        ws.title = "Dados Extraídos"
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplicar formatação ao cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Ajustar largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Formatar células de dados
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        # Rodapé
        footer_row = ws.max_row + 2
        ws[f'A{footer_row}'] = f"Extraído em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Total: {len(df_final)} linhas"
        ws[f'A{footer_row}'].font = Font(italic=True, size=9, color="808080")
        
        wb.save(output_excel)
        
        print(f"✅ Excel formatado salvo!\n")
        
        # 8. Mostrar resumo
        print("=" * 100)
        print("📊 RESUMO DA EXTRAÇÃO")
        print("=" * 100)
        print(f"\n✅ Arquivo Excel: {output_excel}")
        print(f"✅ Total de linhas: {len(df_final)}")
        print(f"✅ Total de colunas: {len(df_final.columns)}")
        print(f"✅ Colunas: {', '.join(df_final.columns[:5])}{'...' if len(df_final.columns) > 5 else ''}")
        
        # Preview dos dados
        print(f"\n📋 PREVIEW (Primeiras 5 linhas):")
        print("-" * 100)
        pd.set_option('display.max_columns', 10)
        pd.set_option('display.width', 100)
        print(df_final.head(5).to_string(index=False))
        
        print("\n" + "=" * 100)
        print("✅ EXTRAÇÃO CONCLUÍDA COM SUCESSO!")
        print("=" * 100)
        
        return True
        
    except Exception as e:
        print(f"\n❌ Erro durante a extração: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    # PDF a ser processado - MUDE AQUI
    pdf_arquivo = "matriz-ES-2025-2.pdf"
    
    # Nome do Excel de saída
    excel_saida = "matriz_extraida.xlsx"
    
    # Executar extração
    sucesso = extrair_pdf_para_excel(pdf_arquivo, excel_saida)
    
    if sucesso:
        print(f"\n✅ Pronto! Abra o arquivo '{excel_saida}' para ver os dados.")
        print(f"💡 Para mudar o PDF, edite a variável 'pdf_arquivo' no código.")

